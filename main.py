from openpyxl import Workbook
from openpyxl.styles import PatternFill
import fnct as hl
import cv2
import numpy as np

path = "img.jpg"
img = cv2.imread(path, 0)
print(img)


# import openpyxl
wb = Workbook()
sheet = wb["Sheet"] # This sheet is created by default

cnt = 0

# projede všechny sloupce
for row in img:
    cnt = cnt + 1
    letrow = hl.nl(cnt)

    print(row)
    # projede sloupec po řádkách
    for line in row:
        print(line)
        # row = row + 1
        color = hl.rgb(0,0,0)
        fill = PatternFill("solid", start_color=color)
        sheet["{}{}".format(letrow, line)].fill = fill  # letter, řádek


wb.save("wb.xlsx")

print("complete")