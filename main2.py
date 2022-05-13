# will work always, but maybe excel wont open, so scale down imagepip install opencv-python



# from openpyxl import Workbook
import openpyxl
# from openpyxl.styles import PatternFill
import fnct as hl
import cv2
import numpy as np

def nothing(x):
    pass
path = "img.jpg"
# path = str(input())
img = cv2.imread(path)
# cv2.namedWindow("Display", cv2.WINDOW_AUTOSIZE)
# cv2.createTrackbar('R','Display',0,255,nothing)

# import openpyxl
wb = openpyxl.Workbook()
# sheet = wb["Sheet"] # This sheet is created by default
sheet = wb.active # This sheet is created by default

cnt = 0
srow, scol = img.shape[:2]   # 1,048,576 rows or 16,384 columns,  řádky, sloupce

# row vodorovně     col vertikálně


# nastavení barev
for row in range(srow):
    print("COLORED", row)
    cnt = 0
    row = row + 1
    for col in range(scol):
        col = col + 1
        cnt = cnt + 1
        # print(col, cnt)
        letcol = hl.nl(cnt)
        color = hl.rgb(img[row - 1][col - 1][2],  img[row - 1][col - 1][1],  img[row - 1][col - 1][0])  #formát RGB, beru z multdim arraye, indexy pozpátku kvůli OpenCV BGR
        fill = openpyxl.styles.PatternFill("solid", start_color=color)
        sheet["{}{}".format(letcol, row)].fill = fill  # letter, řádek

# nastavení šířky na 1
for row in range(srow):
    print("RESIZED ROWS", row)
    row = row + 1
    sheet.row_dimensions[row].height = 3.75
for col in range(scol):
    print("RESIZED COL", col)
    col = col + 1
    coll = hl.nl(col)
    sheet.column_dimensions[coll].width = 0.7



        
wb.save("wb3.xlsx")

print("complete")
