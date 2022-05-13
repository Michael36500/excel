from openpyxl import Workbook
from openpyxl.styles import PatternFill
import fnct as hl
import cv2
import numpy as np

# import openpyxl
wb = Workbook()
sheet = wb["Sheet"] # This sheet is created by default

cnt = 0

size = 255  # row 1048576

# row vodorovně     column vertikálně

for row in range(size):
    cnt = 0
    row = row + 1

    for column in range(size):
        column = column + 1
        cnt = cnt + 1
        # print(column, cnt)
        letcolumn = hl.nl(cnt)
        color = hl.rgb(row,column,0)
        fill = PatternFill("solid", start_color=color)
        sheet["{}{}".format(letcolumn, row)].fill = fill  # letter, řádek


        
wb.save("wb.xlsx")

print("complete")