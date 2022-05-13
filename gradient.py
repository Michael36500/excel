from openpyxl import Workbook
from openpyxl.styles import PatternFill
import fnct as hl
import cv2
import numpy as np

# import openpyxl
wb = Workbook()
sheet = wb["Sheet"] # This sheet is created by default

cnt = 0

srow = 15  # row 1048576 sloupce
scol = 10   # řádky

# row vodorovně     col vertikálně

for row in range(srow):
    cnt = 0
    row = row + 1

    for col in range(scol):
        col = col + 1
        cnt = cnt + 1
        # print(col, cnt)
        letcol = hl.nl(cnt)
        color = hl.rgb(cnt*20,row*20,cnt*20)
        fill = PatternFill("solid", start_color=color)
        sheet["{}{}".format(letcol, row)].fill = fill  # letter, řádek


        
wb.save("wb.xlsx")

print("complete")